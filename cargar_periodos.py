"""
cargar_periodos.py
==================
Carga la tabla periodos a partir de la columna "Periodo" (columna I) de la hoja
"2. Movimientos".

Un período es un mes contable de referencia (ej: 2026-01). No es lo mismo que la
fecha del movimiento: un cobro hecho en enero 2026 puede corresponder al período
diciembre 2025. Por eso los movimientos guardan ambos datos.

Para cada (año, mes) distinto calculamos:
  - fecha_inicio: primer día del mes.
  - fecha_fin:    último día del mes (con calendar.monthrange).
  - descripcion:  "AAAA-MM".

Va ANTES de cargar los movimientos (que apuntan a estos períodos).
Es IDEMPOTENTE.

Correr desde la raíz del proyecto:  py cargar_periodos.py
"""

import calendar
from datetime import date, datetime

from openpyxl import load_workbook

from app.database.conexion import SessionLocal
from app.database.models import Periodo

RUTA_EXCEL = "FLUJO_DE_FONDOS_AUTOS_2026_V3.xlsm"


def leer_periodos(ruta: str):
    """Devuelve un set de tuplas (año, mes) distintas de la columna Periodo."""
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb["2. Movimientos"]

    combos = set()
    for row in ws.iter_rows(min_row=9, values_only=True):
        codigo = str(row[3]).strip() if row[3] is not None else ""
        if codigo in ("", "3000000"):
            continue
        periodo = row[8]  # columna I
        if isinstance(periodo, datetime):
            combos.add((periodo.year, periodo.month))

    wb.close()
    return combos


def main():
    combos = leer_periodos(RUTA_EXCEL)
    print(f"En el Excel: {len(combos)} períodos distintos.")

    creados = 0
    with SessionLocal() as sesion:
        for anio, mes in sorted(combos):
            existe = sesion.query(Periodo).filter_by(mes=mes, anio=anio).one_or_none()
            if existe is not None:
                continue

            ultimo_dia = calendar.monthrange(anio, mes)[1]  # cantidad de días del mes
            sesion.add(
                Periodo(
                    fecha_inicio=date(anio, mes, 1),
                    fecha_fin=date(anio, mes, ultimo_dia),
                    mes=mes,
                    anio=anio,
                    descripcion=f"{anio}-{mes:02d}",
                )
            )
            creados += 1

        sesion.commit()
        print(f"✓ Períodos nuevos: {creados}. Total en la base: {sesion.query(Periodo).count()}")


if __name__ == "__main__":
    main()
