"""
cargar_saldo_inicial.py
=======================
Carga el ÚNICO saldo inicial (el punto de partida del flujo) en la tabla
saldos_iniciales.

De dónde sale el valor:
  En la hoja "2. Movimientos" del Excel, las filas con código 3000000 son las
  de "SALDO INICIAL". Solo la primera (la de fecha más temprana, 2026-01-01)
  tiene un monto real en la columna "Saldo" (columna O). Ese es el número que
  buscamos y guardamos. El resto de las filas 3000000 son marcadores de mes
  sin valor, y las ignoramos.

Es IDEMPOTENTE: si ya existe un saldo inicial para esa fecha, no lo duplica.

Correr desde la raíz del proyecto:  py cargar_saldo_inicial.py
"""

from datetime import datetime
from decimal import Decimal

from openpyxl import load_workbook

from app.database.conexion import SessionLocal
from app.database.models import SaldoInicial

# Ruta al Excel. Ajustala si el archivo está en otra carpeta.
RUTA_EXCEL = "FLUJO_DE_FONDOS_AUTOS_2026_V3.xlsm"

# Código que en el Excel identifica las filas de saldo inicial.
CODIGO_SALDO_INICIAL = "3000000"


def leer_saldo_inicial_del_excel(ruta: str):
    """
    Recorre la hoja de movimientos y devuelve (fecha, monto) del saldo inicial:
    la fila 3000000 más temprana que tenga un monto en la columna Saldo (O).
    Devuelve None si no encuentra ninguna.
    """
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb["2. Movimientos"]

    candidato = None  # tupla (fecha, monto) del más temprano encontrado

    # Los datos arrancan en la fila 9. Columnas (base 0): A=fecha(0),
    # D=codigo(3), O=saldo(14).
    for fila in ws.iter_rows(min_row=9, values_only=True):
        codigo = fila[3]
        saldo = fila[14]
        fecha = fila[0]

        # Nos quedamos solo con las filas de saldo inicial que tengan monto.
        if codigo is None or str(codigo).strip() != CODIGO_SALDO_INICIAL:
            continue
        if saldo is None or fecha is None:
            continue

        # openpyxl devuelve la fecha como datetime; la pasamos a date.
        if isinstance(fecha, datetime):
            fecha = fecha.date()

        # Redondeamos a 2 decimales (el Excel trae ruido de floats).
        monto = Decimal(str(round(float(saldo), 2)))

        # Nos quedamos con la fecha más temprana.
        if candidato is None or fecha < candidato[0]:
            candidato = (fecha, monto)

    wb.close()
    return candidato


def main():
    resultado = leer_saldo_inicial_del_excel(RUTA_EXCEL)

    if resultado is None:
        print("✗ No encontré ninguna fila de saldo inicial con monto en el Excel.")
        return

    fecha, monto = resultado
    print(f"Saldo inicial encontrado en el Excel: {monto:,.2f}  (fecha {fecha})")

    with SessionLocal() as sesion:
        # ¿Ya existe un saldo inicial global (automotriz NULL) para esa fecha?
        existente = (
            sesion.query(SaldoInicial)
            .filter_by(fecha=fecha, automotriz_id=None)
            .one_or_none()
        )

        if existente is not None:
            print("• Ya estaba cargado. No se duplica.")
            return

        nuevo = SaldoInicial(
            fecha=fecha,
            monto=monto,
            automotriz_id=None,  # global
            descripcion="Saldo inicial importado del Excel (FLUJO_DE_FONDOS_AUTOS_2026)",
        )
        sesion.add(nuevo)
        sesion.commit()
        print("✓ Saldo inicial cargado en la tabla saldos_iniciales.")


if __name__ == "__main__":
    main()
