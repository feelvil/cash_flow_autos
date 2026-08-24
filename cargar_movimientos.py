"""
cargar_movimientos.py
=====================
Carga los movimientos históricos (cobros y pagos) de la hoja "2. Movimientos".

Para cada fila con código real:
  - resuelve el plan_cuenta_id a partir del Código,
  - resuelve el periodo_id a partir de la columna Periodo (año/mes),
  - toma fecha, comprobante, ingresos y egresos,
  - calcula neto = ingresos - egresos,
  - atribuye el movimiento al usuario "Sistema" (la carga histórica).

Qué se saltea:
  - filas sin código o el SALDO INICIAL (3000000): no son movimientos,
  - filas estructurales/vacías (sin código y sin monto).

Requiere haber corrido antes, en orden:
    py cargar_datos_base.py
    py cargar_subpartidas_detalles.py
    py cargar_plan_cuentas.py
    py cargar_periodos.py

IDEMPOTENCIA: un movimiento no tiene una "clave única de negocio" (puede haber
dos cobros iguales el mismo día). Para no duplicar las 2742 filas, el script NO
recarga si la tabla ya tiene movimientos. Si querés recargar de cero, primero
vaciá la tabla movimientos.

Correr desde la raíz del proyecto:  py cargar_movimientos.py
"""

from datetime import datetime
from decimal import Decimal

from openpyxl import load_workbook

from app.database.conexion import SessionLocal
from app.database.models import PlanCuenta, Periodo, Usuario, Movimiento

RUTA_EXCEL = "FLUJO_DE_FONDOS_AUTOS_2026_V3.xlsm"
USUARIO_SISTEMA = "Sistema"


def limpio(valor) -> str:
    return str(valor).strip() if valor is not None else ""


def a_decimal(valor) -> Decimal:
    """Convierte un número del Excel a Decimal con 2 decimales. No numérico -> 0."""
    if valor is None or not isinstance(valor, (int, float)):
        return Decimal("0")
    return Decimal(str(round(float(valor), 2)))


def leer_movimientos(ruta: str):
    """Devuelve una lista de dicts, uno por movimiento real."""
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb["2. Movimientos"]

    movs = []
    for row in ws.iter_rows(min_row=9, values_only=True):
        codigo = limpio(row[3])
        if codigo in ("", "3000000"):
            continue

        fecha = row[0]
        if not isinstance(fecha, datetime):
            continue  # sin fecha válida no cargamos el movimiento

        # Período (año, mes) si existe; puede ser None.
        periodo = row[8]
        periodo_am = None
        if isinstance(periodo, datetime):
            periodo_am = (periodo.year, periodo.month)

        movs.append(
            {
                "codigo": int(codigo),
                "fecha": fecha.date(),
                "comprobante": limpio(row[9]) or None,
                "ingresos": a_decimal(row[11]),
                "egresos": a_decimal(row[12]),
                "periodo_am": periodo_am,
            }
        )

    wb.close()
    return movs


def main():
    movs = leer_movimientos(RUTA_EXCEL)
    print(f"En el Excel: {len(movs)} movimientos a cargar.")

    with SessionLocal() as sesion:
        # --- Guardas de prerrequisitos --------------------------------------
        usuario = sesion.query(Usuario).filter_by(nombre=USUARIO_SISTEMA).one_or_none()
        if usuario is None:
            print("✗ No existe el usuario 'Sistema'. Corré: py cargar_datos_base.py")
            return

        # Caches para resolver FKs sin ir a la base por cada movimiento.
        planes = {
            row.codigo: row.id
            for row in sesion.query(PlanCuenta.codigo, PlanCuenta.id).all()
        }
        if not planes:
            print("✗ No hay plan de cuentas. Corré: py cargar_plan_cuentas.py")
            return

        periodos = {(p.anio, p.mes): p.id for p in sesion.query(Periodo).all()}

        # --- Idempotencia: no recargar si ya hay movimientos ----------------
        ya_hay = sesion.query(Movimiento).count()
        if ya_hay > 0:
            print(f"• Ya hay {ya_hay} movimientos cargados. No se recarga (para no duplicar).")
            print("  Si querés recargar de cero, primero vaciá la tabla movimientos.")
            return

        # --- Carga ----------------------------------------------------------
        creados = 0
        salteados = 0
        for m in movs:
            plan_id = planes.get(m["codigo"])
            if plan_id is None:
                print(f"  ⚠ Código {m['codigo']} sin plan de cuentas. Salteado.")
                salteados += 1
                continue

            periodo_id = periodos.get(m["periodo_am"]) if m["periodo_am"] else None

            sesion.add(
                Movimiento(
                    fecha=m["fecha"],
                    mes=m["fecha"].month,
                    plan_cuenta_id=plan_id,
                    periodo_id=periodo_id,
                    comprobante=m["comprobante"],
                    ingresos=m["ingresos"],
                    egresos=m["egresos"],
                    neto=m["ingresos"] - m["egresos"],
                    usuario_id=usuario.id,
                    anulado=False,
                )
            )
            creados += 1

        sesion.commit()

        print(f"✓ Movimientos cargados: {creados} · Salteados: {salteados}")
        print(f"  Total en la base: {sesion.query(Movimiento).count()}")


if __name__ == "__main__":
    main()
