# app/utils/exportar_excel.py
"""
Exportación a Excel (.xlsx) del flujo de fondos.

Este módulo NO calcula nada por su cuenta: le pide los datos ya armados a
reportes.py y solo se ocupa de VOLCARLOS a un Excel prolijo. Así, si el cálculo
del saldo cambia, este archivo no se toca.

Genera dos cosas:
  - exportar_movimientos(...): la planilla tipo "libro", con las MISMAS columnas
    que el Excel original (Fecha, Mes, Día, Código, Partida, Tipo Operación,
    Sub-Partida, Detalle, Período, Comprobante, Automotriz, Ingresos, Egresos,
    Neto, Saldo Acumulado). Es el reemplazo directo de la hoja "2. Movimientos".
  - exportar_resumen(...): un resumen agrupado por la dimensión que se pida
    (partida, sub_partida, etc.), con totales.

Ambas devuelven la ruta (Path) del archivo generado, para que la UI después
pueda abrirlo o mostrar "guardado en ...".

Usa openpyxl (ya está en requirements). Detalle técnico: openpyxl guarda los
números como float igual que Excel, así que convertimos los Decimal a float solo
al escribir la celda. La plata se sigue calculando con Decimal en toda la lógica;
el float es únicamente para la representación en la planilla.
"""

from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database.conexion import SessionLocal
from app.logica.reportes import (
    movimientos_detallados,
    resumen_por_clasificacion,
)


# ---------------------------------------------------------------------------
# Constantes de formato
# ---------------------------------------------------------------------------
# En Excel en español (es-AR) este formato muestra 1.234.567,89: los símbolos
# del código son "placeholders", Excel los pinta con los separadores del sistema.
FORMATO_MONEDA = "#,##0.00"
FORMATO_FECHA = "DD/MM/YYYY"

_AZUL = "1F4E78"        # cabecera
_GRIS_ZEBRA = "F2F2F2"  # filas alternadas

# Definición de columnas de la planilla de movimientos:
#   (encabezado, clave del dict, ancho, tipo de dato)
# El orden de esta lista ES el orden de las columnas en el Excel.
COLUMNAS_MOVIMIENTOS = [
    ("Fecha",           "fecha",           12, "fecha"),
    ("Mes",             "mes",              6, "entero"),
    ("Día",             "dia",              6, "entero"),
    ("Código",          "codigo",          10, "entero"),
    ("Partida",         "partida",         12, "texto"),
    ("Tipo Operación",  "tipo_operacion",  18, "texto"),
    ("Sub-Partida",     "sub_partida",     20, "texto"),
    ("Detalle",         "detalle",         24, "texto"),
    ("Período",         "periodo",         10, "texto"),
    ("Comprobante",     "comprobante",     18, "texto"),
    ("Automotriz",      "automotriz",      24, "texto"),
    ("Ingresos",        "ingresos",        16, "moneda"),
    ("Egresos",         "egresos",         16, "moneda"),
    ("Neto",            "neto",            16, "moneda"),
    ("Saldo Acumulado", "saldo_acumulado", 18, "moneda"),
]


# ---------------------------------------------------------------------------
# Helpers de formato y de ruta
# ---------------------------------------------------------------------------
def _num(valor):
    """Convierte Decimal (o None) a float para escribirlo en la celda."""
    return float(valor) if valor is not None else None


def _aplicar_estilo_encabezado(celda) -> None:
    """Deja una celda con pinta de cabecera: negrita, fondo azul, texto blanco."""
    celda.font = Font(bold=True, color="FFFFFF")
    celda.fill = PatternFill(fill_type="solid", fgColor=_AZUL)
    celda.alignment = Alignment(horizontal="center", vertical="center")


def _resolver_ruta(ruta, prefijo: str) -> Path:
    """
    Decide dónde guardar el archivo:
      - ruta=None  -> nombre por defecto con fecha/hora, en la carpeta actual.
      - ruta=carpeta existente -> nombre por defecto dentro de esa carpeta.
      - ruta=archivo -> se respeta (y se le asegura la extensión .xlsx).
    """
    nombre_default = f"{prefijo}_{datetime.now():%Y%m%d_%H%M}.xlsx"
    if ruta is None:
        return Path.cwd() / nombre_default
    p = Path(ruta)
    if p.is_dir():
        return p / nombre_default
    if p.suffix.lower() != ".xlsx":
        p = p.with_suffix(".xlsx")
    return p


# ---------------------------------------------------------------------------
# Escritura de la hoja de movimientos
# ---------------------------------------------------------------------------
def _escribir_hoja_movimientos(ws, filas: list[dict], titulo: str) -> None:
    """Vuelca las filas de movimientos_detallados() en la hoja 'ws', con formato."""
    n_cols = len(COLUMNAS_MOVIMIENTOS)

    # Fila 1: título (celda combinada a lo ancho de toda la tabla).
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    celda_titulo = ws.cell(row=1, column=1, value=titulo)
    celda_titulo.font = Font(bold=True, size=13)
    celda_titulo.alignment = Alignment(horizontal="left", vertical="center")

    # Fila 2: encabezados + anchos de columna.
    for c, (encabezado, _clave, ancho, _tipo) in enumerate(COLUMNAS_MOVIMIENTOS, start=1):
        celda = ws.cell(row=2, column=c, value=encabezado)
        _aplicar_estilo_encabezado(celda)
        ws.column_dimensions[get_column_letter(c)].width = ancho

    # Filas 3+: datos.
    fila_datos_inicio = 3
    for i, datos in enumerate(filas):
        r = fila_datos_inicio + i
        anulado = datos.get("anulado", False)

        for c, (_enc, clave, _ancho, tipo) in enumerate(COLUMNAS_MOVIMIENTOS, start=1):
            # "Día" no viene en el dict: lo sacamos de la fecha.
            if clave == "dia":
                valor = datos["fecha"].day
            else:
                valor = datos.get(clave)

            celda = ws.cell(row=r, column=c)
            if tipo == "moneda":
                celda.value = _num(valor)
                celda.number_format = FORMATO_MONEDA
            elif tipo == "fecha":
                celda.value = valor
                celda.number_format = FORMATO_FECHA
            else:  # entero o texto: openpyxl los maneja directo
                celda.value = valor

            # Los anulados (si se incluyen) van en gris cursiva para distinguirlos.
            if anulado:
                celda.font = Font(italic=True, color="999999")

        # Zebra: sombreado suave en filas pares, para que sea más fácil de leer.
        if i % 2 == 1 and not anulado:
            relleno = PatternFill(fill_type="solid", fgColor=_GRIS_ZEBRA)
            for c in range(1, n_cols + 1):
                ws.cell(row=r, column=c).fill = relleno

    # Fila de TOTALES (solo si hay datos).
    if filas:
        r_tot = fila_datos_inicio + len(filas)
        total_ingresos = sum((d["ingresos"] for d in filas), Decimal("0"))
        total_egresos = sum((d["egresos"] for d in filas), Decimal("0"))
        saldo_final = filas[-1]["saldo_acumulado"]

        ws.cell(row=r_tot, column=1, value="TOTALES").font = Font(bold=True)
        # Columnas 12=Ingresos, 13=Egresos, 14=Neto, 15=Saldo Acumulado.
        for col, val in [
            (12, total_ingresos),
            (13, total_egresos),
            (14, total_ingresos - total_egresos),
            (15, saldo_final),
        ]:
            celda = ws.cell(row=r_tot, column=col, value=_num(val))
            celda.number_format = FORMATO_MONEDA
            celda.font = Font(bold=True)

        # Línea arriba de la fila de totales, para separarla visualmente.
        borde_arriba = Border(top=Side(style="thin"))
        for col in range(1, n_cols + 1):
            ws.cell(row=r_tot, column=col).border = borde_arriba

        # Autofiltro sobre encabezados + datos (sin incluir la fila de totales).
        ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}{r_tot - 1}"

    # Congelar título + encabezados: al hacer scroll, quedan siempre visibles.
    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
# Función pública: exportar movimientos
# ---------------------------------------------------------------------------
def exportar_movimientos(
    session,
    ruta=None,
    *,
    desde=None,
    hasta=None,
    automotriz_id=None,
    periodo_id=None,
    incluir_anulados=False,
) -> Path:
    """
    Genera un Excel con los movimientos (reproduce la hoja del Excel original).

    Acepta los mismos filtros que reportes.movimientos_detallados. Devuelve la
    ruta (Path) del archivo generado.
    """
    filas = movimientos_detallados(
        session,
        desde=desde,
        hasta=hasta,
        automotriz_id=automotriz_id,
        periodo_id=periodo_id,
        incluir_anulados=incluir_anulados,
    )

    ruta = _resolver_ruta(ruta, "movimientos")
    titulo = (
        f"Flujo de Fondos — Movimientos   "
        f"({len(filas)} filas · generado {datetime.now():%d/%m/%Y %H:%M})"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    _escribir_hoja_movimientos(ws, filas, titulo)
    wb.save(ruta)
    return ruta


# ---------------------------------------------------------------------------
# Función pública: exportar resumen por clasificación
# ---------------------------------------------------------------------------
_COLUMNAS_RESUMEN = [
    ("Cantidad", "cantidad", 12, "entero"),
    ("Ingresos", "ingresos", 18, "moneda"),
    ("Egresos",  "egresos",  18, "moneda"),
    ("Neto",     "neto",     18, "moneda"),
]


def exportar_resumen(
    session,
    dimension: str,
    ruta=None,
    *,
    desde=None,
    hasta=None,
    automotriz_id=None,
    periodo_id=None,
    incluir_anulados=False,
) -> Path:
    """
    Genera un Excel con el resumen agrupado por 'dimension' (partida,
    tipo_operacion, sub_partida, detalle o automotriz), con totales al pie.
    Devuelve la ruta (Path) del archivo.
    """
    filas = resumen_por_clasificacion(
        session,
        dimension,
        desde=desde,
        hasta=hasta,
        automotriz_id=automotriz_id,
        periodo_id=periodo_id,
        incluir_anulados=incluir_anulados,
    )

    ruta = _resolver_ruta(ruta, f"resumen_{dimension}")
    # La primera columna se titula con el nombre de la dimensión, ej: "Sub-Partida".
    etiqueta = dimension.replace("_", " ").title()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Resumen {dimension}"[:31]  # Excel limita el nombre de hoja a 31 chars

    # Título.
    n_cols = 1 + len(_COLUMNAS_RESUMEN)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    celda_titulo = ws.cell(
        row=1, column=1,
        value=f"Resumen por {etiqueta}   (generado {datetime.now():%d/%m/%Y %H:%M})",
    )
    celda_titulo.font = Font(bold=True, size=13)

    # Encabezados: la dimensión + las columnas numéricas.
    encabezados = [etiqueta] + [enc for enc, *_ in _COLUMNAS_RESUMEN]
    anchos = [30] + [ancho for _e, _k, ancho, _t in _COLUMNAS_RESUMEN]
    for c, (enc, ancho) in enumerate(zip(encabezados, anchos), start=1):
        celda = ws.cell(row=2, column=c, value=enc)
        _aplicar_estilo_encabezado(celda)
        ws.column_dimensions[get_column_letter(c)].width = ancho

    # Datos.
    for i, datos in enumerate(filas):
        r = 3 + i
        ws.cell(row=r, column=1, value=datos["nombre"])
        for c, (_enc, clave, _ancho, tipo) in enumerate(_COLUMNAS_RESUMEN, start=2):
            celda = ws.cell(row=r, column=c)
            if tipo == "moneda":
                celda.value = _num(datos[clave])
                celda.number_format = FORMATO_MONEDA
            else:
                celda.value = datos[clave]

    # Totales.
    if filas:
        r_tot = 3 + len(filas)
        ws.cell(row=r_tot, column=1, value="TOTALES").font = Font(bold=True)
        total_ing = sum((d["ingresos"] for d in filas), Decimal("0"))
        total_egr = sum((d["egresos"] for d in filas), Decimal("0"))
        total_cant = sum(d["cantidad"] for d in filas)
        ws.cell(row=r_tot, column=2, value=total_cant).font = Font(bold=True)
        for col, val in [(3, total_ing), (4, total_egr), (5, total_ing - total_egr)]:
            celda = ws.cell(row=r_tot, column=col, value=_num(val))
            celda.number_format = FORMATO_MONEDA
            celda.font = Font(bold=True)
        borde_arriba = Border(top=Side(style="thin"))
        for col in range(1, n_cols + 1):
            ws.cell(row=r_tot, column=col).border = borde_arriba

    ws.freeze_panes = "A3"
    wb.save(ruta)
    return ruta


# ---------------------------------------------------------------------------
# Prueba rápida (genera archivos reales en la carpeta actual)
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    # Correr con:  py -m app.utils.exportar_excel
    # Lee de la base (solo lectura) y ESCRIBE dos .xlsx en la carpeta actual.
    with SessionLocal() as session:
        ruta_mov = exportar_movimientos(session)
        print(f"Movimientos exportados a:\n  {ruta_mov}\n")

        ruta_res = exportar_resumen(session, "sub_partida")
        print(f"Resumen por sub-partida exportado a:\n  {ruta_res}")
