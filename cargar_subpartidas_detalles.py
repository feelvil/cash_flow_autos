"""
cargar_subpartidas_detalles.py
==============================
Carga los niveles 3 y 4 de la jerarquía de clasificación, leyéndolos de la hoja
"4. Plan_cuentas" del Excel:

  - SUB-PARTIDAS: se identifican por la combinación (partida + tipo op. + nombre),
    porque el mismo nombre (ej: "Rentas") se usa en varias combinaciones.
  - DETALLES: cuelgan de una sub-partida puntual, porque el mismo nombre de
    detalle (ej: "BCLP") aparece en más de una sub-partida.

Por eso el orden importa: PRIMERO las sub-partidas, DESPUÉS los detalles.

Requiere que ya estén cargados los datos base (partidas y tipos de operación),
o sea correr antes:  py cargar_datos_base.py

Es IDEMPOTENTE: si algo ya existe, no lo duplica.

Correr desde la raíz del proyecto:  py cargar_subpartidas_detalles.py
"""

from openpyxl import load_workbook

from app.database.conexion import SessionLocal
from app.database.models import Partida, TipoOperacion, SubPartida, Detalle

# Ruta al Excel. Ajustala si el archivo está en otra carpeta.
RUTA_EXCEL = "FLUJO_DE_FONDOS_AUTOS_2026_V3.xlsm"


def limpio(valor) -> str:
    """Convierte a texto y saca espacios de los bordes. None -> ''."""
    return str(valor).strip() if valor is not None else ""


def leer_jerarquia(ruta: str):
    """
    Recorre la hoja "4. Plan_cuentas" y devuelve dos conjuntos:
      - subs: tuplas (partida, tipo, sub)              -> sub-partidas únicas
      - dets: tuplas (partida, tipo, sub, detalle)     -> detalles únicos (con contexto)
    Saltea la fila de SALDO INICIAL (código 3000000) y las filas incompletas.
    """
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb["4. Plan_cuentas"]

    subs = set()
    dets = set()

    # Columnas (base 0): A=codigo(0), B=partida(1), C=tipo(2), D=sub(3), E=detalle(4)
    for row in ws.iter_rows(min_row=2, values_only=True):
        codigo = limpio(row[0])
        if codigo in ("", "3000000"):  # vacías o saldo inicial -> no clasifican
            continue

        partida = limpio(row[1])
        tipo = limpio(row[2])
        sub = limpio(row[3])
        det = limpio(row[4])

        if partida and tipo and sub:
            subs.add((partida, tipo, sub))
        if partida and tipo and sub and det:
            dets.add((partida, tipo, sub, det))

    wb.close()
    return subs, dets


def main():
    subs, dets = leer_jerarquia(RUTA_EXCEL)
    print(f"En el Excel: {len(subs)} sub-partidas y {len(dets)} detalles (contexto completo).")

    creados_sub = 0
    creados_det = 0

    with SessionLocal() as sesion:
        # --- Cacheamos partidas y tipos en diccionarios {nombre: id} ---------
        # Así resolvemos las claves foráneas sin consultar la base una y otra vez.
        partidas = {p.nombre: p.id for p in sesion.query(Partida).all()}
        tipos = {t.nombre: t.id for t in sesion.query(TipoOperacion).all()}

        if not partidas or not tipos:
            print("✗ Faltan partidas o tipos de operación. Corré primero: py cargar_datos_base.py")
            return

        # --- 1) SUB-PARTIDAS ------------------------------------------------
        for partida, tipo, sub in sorted(subs):
            partida_id = partidas.get(partida)
            tipo_id = tipos.get(tipo)
            if partida_id is None or tipo_id is None:
                print(f"  ⚠ Salteada sub-partida {sub!r}: no encontré partida/tipo ({partida}/{tipo}).")
                continue

            existe = (
                sesion.query(SubPartida)
                .filter_by(nombre=sub, partida_id=partida_id, tipo_operacion_id=tipo_id)
                .one_or_none()
            )
            if existe is None:
                sesion.add(
                    SubPartida(nombre=sub, partida_id=partida_id, tipo_operacion_id=tipo_id)
                )
                creados_sub += 1

        # flush(): manda los INSERT de sub-partidas para que tengan id asignado,
        # sin cerrar la transacción todavía. Los necesitamos para los detalles.
        sesion.flush()

        # --- Cacheamos las sub-partidas por su contexto completo -------------
        cache_sub = {
            (sp.nombre, sp.partida_id, sp.tipo_operacion_id): sp.id
            for sp in sesion.query(SubPartida).all()
        }

        # --- 2) DETALLES ----------------------------------------------------
        for partida, tipo, sub, det in sorted(dets):
            partida_id = partidas.get(partida)
            tipo_id = tipos.get(tipo)
            sub_id = cache_sub.get((sub, partida_id, tipo_id))
            if sub_id is None:
                print(f"  ⚠ Salteado detalle {det!r}: no encontré la sub-partida {sub!r}.")
                continue

            existe = (
                sesion.query(Detalle)
                .filter_by(nombre=det, sub_partida_id=sub_id)
                .one_or_none()
            )
            if existe is None:
                sesion.add(Detalle(nombre=det, sub_partida_id=sub_id))
                creados_det += 1

        # Confirmamos todo junto.
        sesion.commit()

        # --- Resumen --------------------------------------------------------
        print(f"✓ Listo. Sub-partidas nuevas: {creados_sub} · Detalles nuevos: {creados_det}")
        print("  Totales en la base:")
        print(f"    Sub-partidas: {sesion.query(SubPartida).count()}")
        print(f"    Detalles:     {sesion.query(Detalle).count()}")


if __name__ == "__main__":
    main()
