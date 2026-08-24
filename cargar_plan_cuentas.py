"""
cargar_plan_cuentas.py
======================
Carga la tabla planes_cuentas: el "hub" que enlaza cada CÓDIGO con toda su
jerarquía (partida, tipo de operación, sub-partida, detalle y automotriz).

Lee la hoja "4. Plan_cuentas" del Excel y, para cada código, resuelve las claves
foráneas contra lo que ya sembramos antes. Por eso este script va ÚLTIMO en la
cadena de carga de clasificación:

    1) py cargar_datos_base.py            (partidas, tipos, automotrices, usuario)
    2) py cargar_subpartidas_detalles.py  (sub-partidas y detalles)
    3) py cargar_plan_cuentas.py          (este)

Detalles importantes del mapeo:
  - Se saltea el código 3000000 (SALDO INICIAL): no se clasifica.
  - Los 492 códigos reales tienen automotriz (ninguno es transversal), pero la
    columna queda nullable por si aparece alguno a futuro.
  - La columna "derechos" está casi vacía (solo 2 filas con valores que parecen
    restos de edición: 'vw' e 'INGRESOS'). Se importa TAL CUAL viene del Excel.
    Si querés limpiarla, avisá y la dejamos siempre en NULL.

Es IDEMPOTENTE: los códigos que ya existan no se reinsertan.

Correr desde la raíz del proyecto:  py cargar_plan_cuentas.py
"""

from openpyxl import load_workbook

from app.database.conexion import SessionLocal
from app.database.models import (
    Partida,
    TipoOperacion,
    SubPartida,
    Detalle,
    Automotriz,
    PlanCuenta,
)

# Ruta al Excel. Ajustala si el archivo está en otra carpeta.
RUTA_EXCEL = "FLUJO_DE_FONDOS_AUTOS_2026_V3.xlsm"


def limpio(valor) -> str:
    """Convierte a texto y saca espacios de los bordes. None -> ''."""
    return str(valor).strip() if valor is not None else ""


def leer_planes(ruta: str):
    """
    Recorre la hoja "4. Plan_cuentas" y devuelve una lista de diccionarios,
    uno por código (salteando vacías y el SALDO INICIAL 3000000).
    Columnas (base 0): A=codigo(0), B=partida(1), C=tipo(2), D=sub(3),
                       E=detalle(4), F=automotriz(5), I=derechos(8).
    """
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb["4. Plan_cuentas"]

    planes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        codigo = limpio(row[0])
        if codigo in ("", "3000000"):
            continue
        planes.append(
            {
                "codigo": int(codigo),
                "partida": limpio(row[1]),
                "tipo": limpio(row[2]),
                "sub": limpio(row[3]),
                "detalle": limpio(row[4]),
                "automotriz": limpio(row[5]),
                # derechos: texto o None si está vacío.
                "derechos": limpio(row[8]) or None,
            }
        )

    wb.close()
    return planes


def main():
    planes = leer_planes(RUTA_EXCEL)
    print(f"En el Excel: {len(planes)} códigos (sin contar SALDO INICIAL).")

    creados = 0
    salteados = 0

    with SessionLocal() as sesion:
        # --- Caches para resolver claves foráneas sin ir a la base por cada código ---
        partidas = {p.nombre: p.id for p in sesion.query(Partida).all()}
        tipos = {t.nombre: t.id for t in sesion.query(TipoOperacion).all()}
        automotrices = {a.nombre: a.id for a in sesion.query(Automotriz).all()}
        # sub-partidas por su contexto completo (nombre + partida + tipo)
        sub_por_ctx = {
            (sp.nombre, sp.partida_id, sp.tipo_operacion_id): sp.id
            for sp in sesion.query(SubPartida).all()
        }
        # detalles por (nombre + sub_partida)
        det_por_ctx = {
            (d.nombre, d.sub_partida_id): d.id for d in sesion.query(Detalle).all()
        }

        if not (partidas and tipos and automotrices and sub_por_ctx and det_por_ctx):
            print("✗ Falta cargar datos base o jerarquía. Corré antes:")
            print("    py cargar_datos_base.py")
            print("    py cargar_subpartidas_detalles.py")
            return

        # Códigos ya cargados, para no reinsertarlos.
        existentes = {row.codigo for row in sesion.query(PlanCuenta.codigo).all()}

        for p in planes:
            if p["codigo"] in existentes:
                continue

            # Resolvemos cada eslabón de la jerarquía.
            partida_id = partidas.get(p["partida"])
            tipo_id = tipos.get(p["tipo"])
            sub_id = sub_por_ctx.get((p["sub"], partida_id, tipo_id))
            det_id = det_por_ctx.get((p["detalle"], sub_id))
            auto_id = automotrices.get(p["automotriz"])  # None si fuera transversal

            # Validación defensiva: si falta un eslabón OBLIGATORIO, salteamos y avisamos.
            faltan = []
            if partida_id is None:
                faltan.append(f"partida {p['partida']!r}")
            if tipo_id is None:
                faltan.append(f"tipo {p['tipo']!r}")
            if sub_id is None:
                faltan.append(f"sub-partida {p['sub']!r}")
            if det_id is None:
                faltan.append(f"detalle {p['detalle']!r}")
            if faltan:
                print(f"  ⚠ Código {p['codigo']} salteado: no resolví {', '.join(faltan)}.")
                salteados += 1
                continue

            sesion.add(
                PlanCuenta(
                    codigo=p["codigo"],
                    partida_id=partida_id,
                    tipo_operacion_id=tipo_id,
                    sub_partida_id=sub_id,
                    detalle_id=det_id,
                    automotriz_id=auto_id,
                    derechos=p["derechos"],
                )
            )
            creados += 1

        sesion.commit()

        print(f"✓ Listo. Planes de cuentas nuevos: {creados} · Salteados: {salteados}")
        print(f"  Total en la base: {sesion.query(PlanCuenta).count()}")


if __name__ == "__main__":
    main()
